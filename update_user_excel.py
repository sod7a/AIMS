import openpyxl
from app import app, User
from datetime import datetime

def update_excel():
    with app.app_context():
        print("Fetching users from database...")
        users = User.query.order_by(User.id.asc()).all()
        
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "User List"
        
        # Headers
        headers = ['ID', 'Username', 'Full Name', 'Department', 'Position', 'Grade', 'Role', 'Email']
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        
        # Add data
        for u in users:
            ws.append([
                u.id,
                u.username,
                u.full_name or '',
                u.department or '',
                u.work_position or '',
                u.working_grade or '',
                u.role,
                u.email_address or ''
            ])
            
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
            
        # Save file
        filename = "USER LIST.xlsx"
        wb.save(filename)
        print(f"Successfully updated {filename} with {len(users)} users.")

if __name__ == "__main__":
    update_excel()
